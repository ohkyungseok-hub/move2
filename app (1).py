import streamlit as st
import pandas as pd
import re
import time
import zipfile
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook

# =========================
# Settings
# =========================
DETAIL_LIMIT = 9
DEFAULT_CHUNK = 100
DEFAULT_OUT_SHEET_INDEX = 0
DEFAULT_ID_COL = "A"  # A파일 상품아이디 기본값(필요시 사이드바에서 변경)

# b 템플릿에서 삭제할 행(엑셀 기준): 1,3,4,5,6 삭제 / 2행(컬럼명) 유지
ROWS_TO_DELETE_1BASED = [1, 3, 4, 5, 6]


# =========================
# Utils
# =========================
def col_idx(col: str) -> int:
    idx = 0
    for c in col.upper():
        idx = idx * 26 + (ord(c) - ord("A") + 1)
    return idx - 1


def uniq_keep_order(seq):
    return list(dict.fromkeys(seq))


def extract_bracket_items(val):
    if pd.isna(val):
        return []
    s = str(val).strip()
    if not s:
        return []

    blocks = re.findall(r"\[([^\]]+)\]", s)
    items = []

    if blocks:
        for blk in blocks:
            blk = blk.strip()
            if not blk:
                continue
            if "," in blk:
                items.extend([p.strip() for p in blk.split(",") if p.strip()])
            else:
                items.append(blk)
    else:
        s2 = s.replace("[", "").replace("]", "").strip()
        if not s2:
            return []
        items = [p.strip() for p in s2.split(",") if p.strip()]

    return items


def build_aw_cell(main_items, detail_items):
    lines = []
    if main_items:
        lines.append(f"main^|^https://m.lastorder.in/{main_items[0]}")
    for i, it in enumerate(detail_items[:DETAIL_LIMIT], start=1):
        lines.append(f"detail_{i}^|^https://m.lastorder.in/{it}")
    return "\n".join(lines)


def validate_a_df(a: pd.DataFrame, id_col_letter: str):
    # 필요한 A컬럼: C,D,E,H,J,M,P,S,T + 상품아이디
    required = ["C", "D", "E", "H", "J", "M", "P", "S", "T", id_col_letter]
    max_needed = max(col_idx(c) for c in required)
    if a.shape[1] <= max_needed:
        missing = [c for c in required if col_idx(c) >= a.shape[1]]
        return False, f"A파일 컬럼 부족: {missing} (현재 컬럼 수: {a.shape[1]})"
    return True, ""


def split_rows(rows: list[dict], chunk_size: int):
    return [rows[i:i + chunk_size] for i in range(0, len(rows), chunk_size)] or [[]]


def get_template_bytes(optional_uploaded):
    """
    1순위: 업로드 템플릿(선택)
    2순위: app(1).py와 같은 폴더의 b.xlsx
    """
    if optional_uploaded is not None:
        return optional_uploaded.getvalue()

    local_path = Path("b.xlsx")
    if local_path.exists():
        return local_path.read_bytes()

    raise FileNotFoundError(
        "b.xlsx 템플릿을 찾을 수 없습니다. "
        "app(1).py와 같은 폴더에 b.xlsx를 두거나, 템플릿을 업로드하세요."
    )


def apply_rows_to_template(template_bytes: bytes, rows: list[dict], sheet_index: int, start_row_after_header: int = 2):
    """
    - 템플릿의 모든 탭(시트) 유지
    - 지정 시트(sheet_index)에서만:
        1) 1,3,4,5,6행 삭제 (2행=컬럼명 유지)
        2) '컬럼명 행' 바로 아래(start_row_after_header)부터 rows 기입

    중요:
    - 행 삭제 후에는 엑셀 행 번호가 당겨지므로,
      여기서는 "컬럼명은 남는다"만 보장하고,
      데이터는 '컬럼명 아래'로 넣기 위해 start_row_after_header=2로 고정합니다.
    """
    wb = load_workbook(BytesIO(template_bytes))
    ws = wb.worksheets[sheet_index]

    # ✅ 큰 번호부터 삭제해야 인덱스 꼬임이 없음
    for r in sorted(ROWS_TO_DELETE_1BASED, reverse=True):
        if r == 2:
            continue
        ws.delete_rows(r, 1)

    # ✅ 컬럼명(원래 2행)이 남아있고, 삭제로 인해 보통 1행이 됨
    # 그래서 "컬럼명 아래"인 2행부터 입력
    start_row = start_row_after_header

    for i, row in enumerate(rows):
        excel_row = start_row + i
        for col_letter, val in row.items():
            ws[f"{col_letter}{excel_row}"] = val

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# =========================
# Core transform (FINAL RULES + NEW CHANGES)
# =========================
def make_b_rows_from_a(a: pd.DataFrame, id_col_letter: str):
    """
    B 매핑:
    - A = 1
    - B = 217089
    - C = 1011307
    - G = A:D (상품명)
    - J = 'n'
    - M = 판매종료일(P) 있으면 2 / 없으면 1 (옵션그룹은 그룹 최소 기준)
    - O = P값, 없으면 2999-12-31 (옵션그룹은 그룹 최소)
    - S = A:H
    - T = (A:H - A:J) 결과 + "-1"
    - W = (비옵션 재고) A:M
    - AG = (옵션재고) A:M 을 옵션값 순서대로 ^|^
    - AP = 유효일자(날짜만): P가 없으면 빈칸("")
    - AU = 항상 빈칸("")
    - AR, AS = A:C (C가 없으면 빈칸)
    - AW = 이미지(main + detail_1~9 줄바꿈) 옵션그룹은 전체 합쳐 중복 제거

    옵션그룹:
    - AB='y', AC='선택', AD=옵션값(E)^|^
    """

    pid = a.iloc[:, col_idx(id_col_letter)].astype(str).fillna("").str.strip()
    is_dup = pid.duplicated(keep=False)
    option_pids = set(pid[is_dup])

    # 대표행(각 pid 첫 행)
    rep_mask = ~pid.duplicated(keep="first")
    a_rep = a.loc[rep_mask].reset_index(drop=True)
    pid_rep = pid.loc[rep_mask].reset_index(drop=True)

    # 그룹 최소 판매종료일
    p_all_dt = pd.to_datetime(a.iloc[:, col_idx("P")], errors="coerce")
    p_min_map = p_all_dt.groupby(pid).min()

    # 옵션값(E) -> AD
    e_series = a.iloc[:, col_idx("E")]
    opt_value_map = (
        pd.DataFrame({"pid": pid, "opt": e_series})
        .groupby("pid", sort=False)["opt"]
        .apply(lambda s: "^|^".join(
            uniq_keep_order([str(v).strip() for v in s.tolist() if pd.notna(v) and str(v).strip()])
        ))
        .to_dict()
    )

    # 옵션재고(M) -> AG (옵션값 순서에 맞춰 ^|^)
    m_stock_series = a.iloc[:, col_idx("M")]
    df_opt = pd.DataFrame({"pid": pid, "opt": e_series, "stk": m_stock_series})

    opt_stock_map = {}
    for pid_val, grp in df_opt.groupby("pid", sort=False):
        opt_vals_raw = [str(v).strip() for v in grp["opt"].tolist() if pd.notna(v) and str(v).strip()]
        opt_vals = uniq_keep_order(opt_vals_raw)

        stocks_out = []
        for ov in opt_vals:
            sub = grp.loc[grp["opt"].astype(str).str.strip() == ov, "stk"]
            chosen = ""
            for sv in sub.tolist():
                if pd.isna(sv):
                    continue
                ss = str(sv).strip()
                if ss and ss.lower() != "nan":
                    chosen = ss
                    break
            stocks_out.append(chosen)

        opt_stock_map[pid_val] = "^|^".join(stocks_out)

    # 이미지 그룹 합치기
    s_img = a.iloc[:, col_idx("S")]
    t_img = a.iloc[:, col_idx("T")]

    def group_images(pid_value: str):
        mask = (pid == pid_value).to_numpy()

        main_candidates = []
        for sv in s_img[mask]:
            main_candidates.extend(extract_bracket_items(sv))
        main_candidates = [x for x in main_candidates if x]

        detail_candidates = []
        for tv in t_img[mask]:
            detail_candidates.extend(extract_bracket_items(tv))
        detail_candidates = uniq_keep_order([x for x in detail_candidates if x])

        return main_candidates, detail_candidates

    # 계산용 (대표행)
    h_num = pd.to_numeric(a_rep.iloc[:, col_idx("H")], errors="coerce").fillna(0).to_numpy()
    j_num = pd.to_numeric(a_rep.iloc[:, col_idx("J")], errors="coerce").fillna(0).to_numpy()

    out_rows = []
    for i in range(len(a_rep)):
        pid_i = pid_rep.iloc[i]
        is_option = pid_i in option_pids

        row = {}
        row["A"] = 1
        row["B"] = 217089
        row["C"] = 1011307
        row["J"] = "n"

        # 상품명
        row["G"] = a_rep.iloc[:, col_idx("D")].to_numpy()[i]

        # S = A:H
        row["S"] = a_rep.iloc[:, col_idx("H")].to_numpy()[i]

        # T = (H - J) + "-1"
        row["T"] = f"{int(h_num[i] - j_num[i])}-1"

        # AU = 항상 빈칸
        row["AU"] = ""

        # AR/AS = A:C (없으면 빈칸)
        c_raw = a_rep.iloc[:, col_idx("C")].to_numpy()[i]
        c_str = "" if pd.isna(c_raw) or str(c_raw).strip().lower() == "nan" or str(c_raw).strip() == "" else c_raw
        row["AR"] = c_str
        row["AS"] = c_str

        # 판매종료일: 그룹 최소값
        pmin = p_min_map.get(pid_i, pd.NaT)
        if pd.isna(pmin):
            row["M"] = 1
            row["O"] = "2999-12-31"   # O는 기존 규칙 유지
            row["AP"] = ""            # ✅ 변경: P 없으면 AP는 빈칸
        else:
            d = pd.Timestamp(pmin).strftime("%Y-%m-%d")
            row["M"] = 2
            row["O"] = d
            row["AP"] = d            # 날짜만

        # AW 이미지
        main_items, detail_items = group_images(pid_i)
        row["AW"] = build_aw_cell(main_items, detail_items)

        # ✅ 비옵션 재고: W = A:M
        if not is_option:
            w_raw = a_rep.iloc[:, col_idx("M")].to_numpy()[i]
            row["W"] = "" if pd.isna(w_raw) else w_raw

        # 옵션 처리
        if is_option:
            row["AB"] = "y"
            row["AC"] = "선택"
            row["AD"] = opt_value_map.get(pid_i, "")
            row["AG"] = opt_stock_map.get(pid_i, "")

        out_rows.append(row)

    return out_rows


# =========================
# Streamlit UI
# =========================
st.set_page_config(page_title="A→B 변환기(최종)", layout="wide")
st.title("📦 A파일 → B템플릿(b.xlsx) 자동 변환기 (최종)")

with st.expander("동작 요약", expanded=True):
    st.write(
        "- 템플릿 업로드 없이도 **같은 폴더의 b.xlsx를 자동 사용**합니다.\n"
        "- 출력 파일(지정 시트)에서만 **1,3,4,5,6행 삭제**하고 **2행(컬럼명)은 유지**합니다.\n"
        "- **AP는 P가 없으면 빈칸**, **AU는 항상 빈칸**, **C가 없으면 AR/AS는 빈칸**입니다.\n"
        "- 비옵션은 **W=재고**, 옵션은 **AG=옵션재고(^|^)** 입니다.\n"
    )

st.sidebar.header("설정")
id_col_letter = st.sidebar.text_input("A파일 상품아이디 컬럼(엑셀 문자)", value=DEFAULT_ID_COL).strip().upper()
chunk_size = st.sidebar.number_input("분할 저장(행)", min_value=10, max_value=5000, value=DEFAULT_CHUNK, step=10)
sheet_index = st.sidebar.number_input("템플릿에 쓸 시트 인덱스(0=첫 시트)", min_value=0, max_value=30, value=DEFAULT_OUT_SHEET_INDEX, step=1)

template_file = st.file_uploader("B 템플릿 업로드(선택)", type=["xlsx"])
a_files = st.file_uploader("A파일 업로드(여러 개 가능)", type=["xlsx"], accept_multiple_files=True)

run_btn = st.button("🚀 변환 시작", disabled=not a_files)

if run_btn:
    t0 = time.time()
    st.info("처리 중...")

    try:
        template_bytes = get_template_bytes(template_file)
    except Exception as e:
        st.error(str(e))
        st.stop()

    summary_rows = []
    error_rows = []

    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for uf in a_files:
            if uf.name.startswith("~$"):
                continue

            started = time.time()
            status = "OK"
            msg = ""
            input_rows = 0
            out_files = 0

            try:
                a_df = pd.read_excel(uf)
                input_rows = len(a_df)

                ok, vmsg = validate_a_df(a_df, id_col_letter)
                if not ok:
                    raise ValueError(vmsg)

                rows = make_b_rows_from_a(a_df, id_col_letter)
                chunks = split_rows(rows, int(chunk_size))

                for idx, chunk in enumerate(chunks, start=1):
                    out_xlsx = apply_rows_to_template(
                        template_bytes=template_bytes,
                        rows=chunk,
                        sheet_index=int(sheet_index),
                        start_row_after_header=2  # 컬럼명 아래로 입력
                    )
                    out_name = f"{Path(uf.name).stem}_part{idx:03d}.xlsx"
                    zf.writestr(out_name, out_xlsx)
                    out_files += 1

            except Exception as e:
                status = "FAIL"
                msg = str(e)
                error_rows.append({"file": uf.name, "reason": msg})

            elapsed = round(time.time() - started, 3)
            summary_rows.append({
                "file": uf.name,
                "status": status,
                "input_rows": input_rows,
                "output_files": out_files,
                "seconds": elapsed,
                "message": msg
            })

        summary_df = pd.DataFrame(summary_rows)
        zf.writestr("summary_report.csv", summary_df.to_csv(index=False).encode("utf-8-sig"))

        if error_rows:
            errors_df = pd.DataFrame(error_rows)
            zf.writestr("errors.csv", errors_df.to_csv(index=False).encode("utf-8-sig"))

    zip_buf.seek(0)
    total_sec = round(time.time() - t0, 2)

    st.success(f"✅ 완료! 총 소요 {total_sec}s")
    st.subheader("📊 요약 리포트")
    st.dataframe(pd.DataFrame(summary_rows), use_container_width=True)

    if error_rows:
        st.subheader("⚠️ 에러")
        st.dataframe(pd.DataFrame(error_rows), use_container_width=True)

    st.download_button(
        "📦 결과 ZIP 다운로드",
        data=zip_buf,
        file_name="B_result.zip",
        mime="application/zip"
    )
